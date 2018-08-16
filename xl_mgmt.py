import openpyxl
import pandas as pd

from akpy.stnd_vars import StandardMessages
from akpy.time_mgmt import *


def verify_expected_sheet_names(dirpath_xl_file, list_expected_sheet_names):
    missing_sheetnames = []
    wb = openpyxl.load_workbook(dirpath_xl_file, read_only=True)
    for sheet_name in list_expected_sheet_names:
        if sheet_name not in wb.sheetnames:
            missing_sheetnames.append(sheet_name)

    if len(missing_sheetnames) > 0:
        print('[MISSING SHEETS] the selected file is missing the following sheet(s):',
              missing_sheetnames)
        sys.exit('[TERMINATING PROCESS] ' + StandardMessages.msg_contactak)
    else:
        print('[', time_stamp(), ']', 'all expected sheets were verified')


def verify_expected_columns(dirpath_xl_file, target_sheet_name, list_expected_column_names, skiprows=0):
    missing_columns = []
    df = pd.read_excel(dirpath_xl_file, sheet_name=target_sheet_name, skiprows=skiprows)
    columns = list(df.columns)
    for expected_column in list_expected_column_names:
        if expected_column not in columns:
            missing_columns.append(expected_column)

    if len(missing_columns) > 0:
        print('[MISSING COLUMNS]', target_sheet_name, 'tab in the selected file is missing the following column(s):',
              missing_columns)
        sys.exit('[TERMINATING PROCESS] ' + StandardMessages.msg_contactak)
    else:
        print('[', time_stamp(), ']', 'all expected columns were verified')


if __name__ == '__main__':
    # print(time_stamp())
    pass
